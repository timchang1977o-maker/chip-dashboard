#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""美股財報區 — 兩個獨立來源合成一頁 us_earnings.html（chip-dashboard / GitHub Pages）。

來源 A｜韭菜王 Allen：投資 cover list（Google Sheet，link-可讀免登入）→ openpyxl。
  欄位對應：B-E 歷史(A)｜F=本季Allen G=共識 H=實際｜I=下季Allen J=下季共識 K=指引｜L/M=下下季
來源 B｜Henry(Pentimetrics)：curated henry_data.json（含 stance；由 pentimetrics_db 精準抽取）

版面：精煉財經編輯風、強制淺色。**依多空立場分組**；每張卡兩欄——
  左＝最新季（實際 vs 共識的預期差＋比較條），右＝下季展望（Allen vs 共識預期差＋公司指引）。

用法：python3 build_us_earnings.py            （自動下載最新 sheet）
      python3 build_us_earnings.py local.xlsx  （用本機 xlsx，離線）
"""
import sys, json, io, re, datetime, html, urllib.request
import openpyxl
from pathlib import Path

SHEET_ID = "1Hfb1eX23xCbGMYOh-76_DjpAgwl7-8DG8TjbtxqZRD4"
EXPORT = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
UA = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"
ROOT = Path(__file__).resolve().parent

META = {
    "GOOG": ("Alphabet", "看多", "2026/07/23"), "AMZN": ("亞馬遜", "看多", "2026/08/03"),
    "MSFT": ("微軟", "中性偏多", "2026/07/31"), "META": ("Meta", "偏空", "2026/08/02"),
    "TSM": ("台積電 ADR", "中性", "2026/07/17"), "MTK": ("聯發科", "看多", "2026/08/02"),
    "ASML": ("艾司摩爾", "看多", "2026/07/15"), "AMD": ("超微 AMD", "看多", "2026/08/06"),
    "TER": ("Teradyne", "看多", "2026/08/04"), "VRT": ("Vertiv", "中性偏多", "2026/08/09"),
    "ONTO": ("Onto", "看多", "2026/08/13"), "LITE": ("Lumentum", "看多", "2026/08/14"),
    "CLS": ("Celestica", "看多", "2026/07/29"), "GLW": ("康寧 Corning", "中性", "2026/07/30"),
    "NOK": ("Nokia", "看多", "2026/07/27"), "COHR": ("Coherent", "看多", "2026/05/10"),
    "AMAT": ("應用材料", "看多", "2026/05/16"), "AVGO": ("博通 Broadcom", "中性", "2026/07/23"),
    "NVDA": ("輝達 NVIDIA", "看多", "2026/05/28"), "CIEN": ("Ciena", "追蹤", "2026/03/12"),
}
ORDER = list(META.keys())
COL = {c: i for i, c in enumerate("BCDEFGHIJKLM", start=2)}
# 立場顯示順序＋樣式
STANCES = ["看多", "中性偏多", "中性", "中性偏空", "偏空", "追蹤"]
STANCE_CLS = {"看多": "s-bull", "中性偏多": "s-bullmid", "中性": "s-neu",
              "中性偏空": "s-bearmid", "偏空": "s-bear", "追蹤": "s-watch"}


def load_wb():
    if len(sys.argv) > 1:
        p = sys.argv[1]
        return (openpyxl.load_workbook(p, data_only=True),
                openpyxl.load_workbook(p, data_only=False))
    req = urllib.request.Request(EXPORT, headers={"User-Agent": UA})
    raw = urllib.request.urlopen(req, timeout=60).read()
    return (openpyxl.load_workbook(io.BytesIO(raw), data_only=True),
            openpyxl.load_workbook(io.BytesIO(raw), data_only=False))


def num(v):
    return v if isinstance(v, (int, float)) else None


def extract(wbd, wbf):
    data = {}
    for t in ORDER:
        if t not in wbd.sheetnames:
            continue
        wd, wf = wbd[t], wbf[t]
        c = lambda col, row: wd.cell(row=row, column=COL[col]).value
        lab = lambda col: (str(wf.cell(row=2, column=COL[col]).value).strip()
                           if wf.cell(row=2, column=COL[col]).value else "")
        unit = (wf["A2"].value or "").replace("In Millions of", "").strip().upper()
        rec = {
            "name": META[t][0], "stance": META[t][1], "date": META[t][2],
            "unit": unit, "cur_q": lab("F").replace("(E)", ""),
            "act_rev": num(c("H", 5)), "cons_rev": num(c("G", 5)), "est_rev": num(c("F", 5)),
            "act_eps": num(c("H", 9)), "cons_eps": num(c("G", 9)), "est_eps": num(c("F", 9)),
            "n_q": lab("I").replace("(E)", ""),
            "n_est_rev": num(c("I", 5)), "n_est_eps": num(c("I", 9)),
            "n_cons_rev": num(c("J", 5)), "n_cons_eps": num(c("J", 9)),
            "guid": c("K", 5),
            "eps26": num(c("F", 17)), "eps27": num(c("H", 17)),
        }
        price = None
        for r in (19, 20, 21):
            for cc in ("B", "C"):
                f = wf.cell(row=r, column=COL[cc]).value
                if isinstance(f, str) and "GOOGLEFINANCE" in f and "252" not in f:
                    price = num(wd.cell(row=r, column=COL[cc]).value)
                    break
            if price:
                break
        rec["price"] = price
        pe = num(wd["C19"].value)
        if not (pe and 0 < pe < 200):
            pe = (price / rec["eps26"]) if (price and rec["eps26"] and 0 < rec["eps26"] < 500) else None
        rec["pe"] = pe
        data[t] = rec
    return data


# ---------- helpers ----------
def is_eps(v):
    return isinstance(v, (int, float)) and abs(v) < 500


def bil(v):
    return f"{v/100:,.1f}" if isinstance(v, (int, float)) else None


def fmt(v, d=2):
    return f"{v:,.{d}f}" if isinstance(v, (int, float)) else None


def pct(a, b):
    if not (isinstance(a, (int, float)) and isinstance(b, (int, float)) and b):
        return None
    return (a / b - 1) * 100


def esc(v):
    return html.escape(str(v)) if v is not None else ""


def pnum(s):
    """從 Henry 字串取第一個數字（"91.2億美元"→91.2）。"""
    if not isinstance(s, str):
        return None
    m = re.search(r"-?\d+(?:\.\d+)?", s.replace(",", ""))
    return float(m.group()) if m else None


def add_commas(s):
    """把字串裡 4 位以上的整數加千分位（用於 Allen 指引的純數字區間，如 152200~159800）。"""
    if not isinstance(s, str):
        return s
    return re.sub(r"\d{4,}", lambda m: f"{int(m.group()):,}", s)


def diff_row(p, cons_str):
    """預期差列：pill(▲/▼ %) + vs 共識文字（比較條已移除，pill 即表達差距）。"""
    pill = ""
    if p is not None:
        cls = "beat" if p >= 0 else "miss"
        arr = "▲" if p >= 0 else "▼"
        pill = f'<span class="pill {cls}">{arr}{abs(p):.1f}%</span>'
    cl = f'<div class="cons">{cons_str}</div>' if cons_str else ""  # cons_str 已是安全 HTML
    if not (pill or cl):
        return ""
    return f'<div class="diff">{pill}</div>{cl}' if pill else cl


def metric(label, big, unit, p, cons_str, big_cls=""):
    if big is None:
        return (f'<div class="metric"><div class="mlabel">{esc(label)}</div>'
                f'<div class="mbig mut {big_cls}">待公布</div></div>')
    u = f'<span class="munit">億{unit}</span>' if unit else ""
    return (f'<div class="metric"><div class="mlabel">{esc(label)}</div>'
            f'<div class="mbig {big_cls}">{esc(big)}{u}</div>{diff_row(p, cons_str)}</div>')


# ---------- Allen card ----------
def allen_card(t, r):
    scls = STANCE_CLS.get(r["stance"], "s-neu")
    unit = r["unit"] or "USD"
    # 左欄：最新季（實際為主，並列 Allen 估／BBG 共識）
    def cmp3(est, cons):
        segs = []
        if est is not None: segs.append(f'Allen 估 {est}')
        if cons is not None: segs.append(f'<b>共識 {cons}</b>')
        return " · ".join(segs)
    rev_m = metric("實際營收", bil(r["act_rev"]), unit, pct(r["act_rev"], r["cons_rev"]),
                   cmp3(bil(r["est_rev"]), bil(r["cons_rev"])))
    eps_m = ""
    if is_eps(r["act_eps"]):
        eps_m = metric("EPS", fmt(r["act_eps"]), "", pct(r["act_eps"], r["cons_eps"]),
                       cmp3(fmt(r["est_eps"]) if is_eps(r["est_eps"]) else None,
                            fmt(r["cons_eps"]) if is_eps(r["cons_eps"]) else None), "sm")
    left = (f'<div class="col"><div class="col-h">最新季 · {esc(r["cur_q"] or "—")}'
            f'<span class="reported">已公布</span></div>{rev_m}{eps_m}</div>')
    # 右欄：下季展望（Allen vs 共識預期差）+ 公司指引
    nq = r["n_q"] or ""
    nrev = bil(r["n_est_rev"]); neps = fmt(r["n_est_eps"]) if is_eps(r["n_est_eps"]) else None
    nrev_m = metric("Allen 營收", nrev, unit, pct(r["n_est_rev"], r["n_cons_rev"]),
                    f'<b>共識 {bil(r["n_cons_rev"])}</b>' if bil(r["n_cons_rev"]) else "") if nrev else ""
    neps_m = metric("Allen EPS", neps, "", pct(r["n_est_eps"], r["n_cons_eps"]),
                    f'<b>共識 {fmt(r["n_cons_eps"])}</b>' if is_eps(r["n_cons_eps"]) else "", "sm") if neps else ""
    guid = add_commas(r["guid"]) if isinstance(r["guid"], str) else (bil(r["guid"]) if isinstance(r["guid"], (int, float)) else None)
    guid_m = f'<div class="guidbox"><span class="gl">公司指引</span><span class="gv">{esc(guid)}</span></div>' if guid else ""
    right = ""
    if nrev_m or neps_m or guid_m:
        right = (f'<div class="col col-fwd"><div class="col-h">下季展望 · {esc(nq)}</div>'
                 f'{nrev_m}{neps_m}{guid_m}</div>')
    stats = [("FY26", fmt(r["eps26"]) if is_eps(r["eps26"]) else "—"),
             ("FY27", fmt(r["eps27"]) if is_eps(r["eps27"]) else "—"),
             ("股價", fmt(r["price"]) if r["price"] else "—"),
             ("PE", fmt(r["pe"], 1) if r["pe"] else "—")]
    stat_html = "".join(f'<div class="stat"><div class="sl">{k}</div><div class="sv">{v}</div></div>' for k, v in stats)
    searchtxt = f"{t} {r['name']}".lower()
    return f"""<article class="card {scls}" data-t="{t}" data-stance="{r['stance']}" data-s="{searchtxt}">
 <header class="chead"><div class="ident"><span class="sym">{t}</span><span class="nm">{esc(r['name'])}</span></div>
   <span class="badge {scls}">{r['stance']}</span></header>
 <div class="cols">{left}{right}</div>
 <div class="stats">{stat_html}</div>
</article>"""


# ---------- Henry card ----------
def henry_card(h):
    t = h.get("ticker", "?")
    stance = h.get("stance") or "追蹤"
    scls = STANCE_CLS.get(stance, "s-watch")
    # 左欄：最新季（字串金額，解析出預期差）
    rev = h.get("actual_rev"); cr = h.get("cons_rev")
    prev = None
    if isinstance(rev, str) and isinstance(cr, str) and "億" in rev and "億" in cr and "兆" not in rev:
        prev = pct(pnum(rev), pnum(cr))
    vs = []
    if cr: vs.append(f"共識 {cr}")
    if h.get("buyside_rev"): vs.append(f'buyside {h["buyside_rev"]}')
    rev_m = (f'<div class="metric"><div class="mlabel">營收</div>'
             f'<div class="mbig">{esc(rev or "—")}</div>'
             f'{diff_row(prev, "vs " + " · ".join(esc(x) for x in vs) if vs else "")}</div>') if (rev or cr) else ""
    eps = h.get("actual_eps"); ce = h.get("cons_eps")
    peps = pct(pnum(eps), pnum(ce)) if (eps and ce) else None
    eps_m = (f'<div class="metric"><div class="mlabel">EPS</div>'
             f'<div class="mbig sm">{esc(eps or "—")}</div>'
             f'{diff_row(peps, "vs 共識 " + esc(ce) if ce else "")}</div>') if (eps or ce) else ""
    left = f'<div class="col"><div class="col-h">已公布 · {esc(h.get("quarter","") or "—")}</div>{rev_m}{eps_m}</div>'
    # 右欄：展望／指引
    guide = h.get("guide")
    right = (f'<div class="col col-fwd"><div class="col-h">展望／指引</div>'
             f'<div class="gd">{esc(guide)}</div></div>') if guide else ""
    view = f'<blockquote class="cview">{esc(h.get("view"))}</blockquote>' if h.get("view") else ""
    src = " · ".join(x for x in [esc(h.get("src_label", "")), esc(h.get("src_date", ""))] if x)
    searchtxt = f'{t} {h.get("company","")}'.lower()
    return f"""<article class="card {scls}" data-t="{t}" data-stance="{esc(stance)}" data-s="{searchtxt}">
 <header class="chead"><div class="ident"><span class="sym">{t}</span><span class="nm">{esc(h.get('company',''))}</span></div>
   <span class="badge {scls}">{esc(stance)}</span></header>
 <div class="qtag src">{src}</div>
 <div class="cols">{left}{right}</div>
 {view}
</article>"""


def stance_groups(cards_by_stance):
    """把 {stance: [card_html,...]} 依 STANCES 順序渲染成分組帶。"""
    out = []
    for st in STANCES:
        cards = cards_by_stance.get(st)
        if not cards:
            continue
        scls = STANCE_CLS.get(st, "s-neu")
        out.append(
            f'<div class="stance-group" data-stance="{st}">'
            f'<div class="sband {scls}"><span class="sdot"></span>{st}<em>{len(cards)}</em></div>'
            f'<div class="grid">{"".join(cards)}</div></div>')
    return "\n".join(out)


NAV = ('<nav class="topnav"><a href="index.html">📊 籌碼總覽</a>'
       '<a href="institutions.html">🏦 法人買賣超</a>'
       '<a href="us_earnings.html" class="on">📈 美股財報</a></nav>')


def build_html(data, henry, updated):
    # Allen 依立場分組
    a_by = {}
    for t in ORDER:
        if t in data:
            a_by.setdefault(data[t]["stance"], []).append(allen_card(t, data[t]))
    allen_groups = stance_groups(a_by)
    # Henry 依立場分組
    h_by = {}
    for h in henry:
        h_by.setdefault(h.get("stance") or "追蹤", []).append(henry_card(h))
    henry_groups = stance_groups(h_by)

    n = len(data)
    bull = sum(1 for r in data.values() if r["stance"] == "看多")
    bear = sum(1 for r in data.values() if r["stance"] == "偏空")
    ALIAS = {"GOOGL": "GOOG", "TSMC": "TSM"}
    henry_new = [h["ticker"] for h in henry
                 if ALIAS.get(h.get("ticker"), h.get("ticker")) not in ORDER] if henry else []
    henry_sec = ""
    if henry:
        asof = max((h.get("src_date", "") for h in henry), default="")
        asof_txt = f'・資料截至 <b>{esc(asof)}</b>' if asof else ""
        henry_sec = f"""
<section class="src-block henry-block">
<div class="secthead"><h2><span class="dot">H</span>Henry · Pentimetrics</h2>
<span class="secsub">The Trace 券商彙整・市場（Bloomberg）共識與 buy side・依 Henry 評語分立場。{len(henry)} 檔，其中 <b>{len(henry_new)}</b> 檔為 cover list 未收{asof_txt}。</span></div>
{henry_groups}
</section>"""
    return f"""<!doctype html><html lang="zh-Hant"><head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1,viewport-fit=cover">
<title>美股財報區 · Cover List</title>
<link rel="manifest" href="manifest.webmanifest">
<meta name="theme-color" content="#f4f1ea">
<meta name="color-scheme" content="light">
<meta name="apple-mobile-web-app-capable" content="yes">
<meta name="mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-title" content="美股財報">
<link rel="apple-touch-icon" href="icon-180.png">
<link rel="icon" type="image/png" sizes="192x192" href="icon-192.png">
<link rel="icon" href="favicon.ico">
<style>
:root{{
  color-scheme:light;
  --bg:#f4f1ea; --bg2:#efeae0; --panel:#fffdf9; --ink:#232019; --ink2:#4a453b;
  --muted:#8f887a; --hair:#e7e1d5; --line:#ded7c8;
  --accent:#1a5e4a; --accent2:#c2410c;
  --beat:#1f7a4d; --beat-bg:#e6f2ea; --miss:#c0392b; --miss-bg:#fbeae7;
  --gd:#8a6d1f; --henry:#6d4bb8; --henry-bg:#efeaf8;
  --bull:#1f7a4d; --bear:#c0392b; --neu:#8f887a; --bullmid:#2563b8; --bearmid:#c2410c;
  --shadow:0 1px 2px rgba(60,50,30,.04),0 8px 24px rgba(60,50,30,.06);
  --serif:"Iowan Old Style","Palatino Linotype",Palatino,"Songti TC","Noto Serif TC",Georgia,serif;
  --sans:-apple-system,BlinkMacSystemFont,"Segoe UI","PingFang TC","Noto Sans TC",sans-serif;
  --mono:"SF Mono",ui-monospace,"Roboto Mono","DejaVu Sans Mono",Menlo,monospace;
}}
*{{box-sizing:border-box}}
html{{color-scheme:light}}
body{{margin:0;background:var(--bg);color:var(--ink);font-family:var(--sans);
  -webkit-font-smoothing:antialiased;padding:20px 16px 56px;
  background-image:radial-gradient(circle at 1px 1px,rgba(120,105,75,.05) 1px,transparent 0);background-size:22px 22px}}
.wrap{{max-width:1280px;margin:0 auto}}
.topnav{{display:flex;gap:7px;margin:0 0 18px;flex-wrap:wrap}}
.topnav a{{font:600 13px var(--sans);color:var(--muted);text-decoration:none;padding:8px 15px;
  border-radius:11px;border:1px solid var(--hair);background:var(--panel);letter-spacing:.2px;box-shadow:var(--shadow)}}
.topnav a.on{{background:var(--ink);color:var(--bg);border-color:var(--ink)}}
.masthead{{margin:0 0 6px;padding-bottom:16px;border-bottom:2.5px solid var(--ink)}}
h1{{font-family:var(--serif);font-size:31px;font-weight:600;margin:0 0 6px;letter-spacing:.2px;line-height:1.1}}
.lede{{color:var(--ink2);font-size:13px;margin:0;line-height:1.7;max-width:82ch}}
.lede b{{color:var(--ink);font-weight:700}}
.updated{{color:var(--muted);font-size:11.5px;margin-top:7px;font-family:var(--mono);letter-spacing:.3px}}
.kpis{{display:flex;gap:0;flex-wrap:wrap;margin:16px 0 6px;border:1px solid var(--hair);
  border-radius:14px;overflow:hidden;background:var(--panel);box-shadow:var(--shadow)}}
.kpi{{flex:1;min-width:120px;padding:13px 18px;border-right:1px solid var(--hair)}}
.kpi:last-child{{border-right:0}}
.kpi .v{{font-family:var(--serif);font-size:26px;font-weight:600;line-height:1}}
.kpi .l{{font-size:11px;color:var(--muted);margin-top:5px;letter-spacing:.3px}}
.bar{{display:flex;gap:8px;flex-wrap:wrap;align-items:center;margin:20px 0 6px}}
.bar input{{font:500 13px var(--sans);color:var(--ink);background:var(--panel);
  border:1px solid var(--line);border-radius:10px;padding:9px 13px;box-shadow:var(--shadow);min-width:210px}}
.bar .chip{{cursor:pointer;color:var(--muted);border:1px solid var(--line);background:var(--panel);
  border-radius:20px;padding:7px 14px;font:600 12px var(--sans);box-shadow:var(--shadow);transition:all .12s}}
.bar .chip.on{{background:var(--ink);color:var(--bg);border-color:var(--ink)}}
.secthead{{display:flex;align-items:baseline;gap:14px;flex-wrap:wrap;margin:30px 0 14px}}
.secthead h2{{font-family:var(--serif);font-size:20px;font-weight:600;margin:0;letter-spacing:.2px;display:flex;align-items:center;gap:10px}}
.secthead .dot{{display:inline-grid;place-items:center;width:26px;height:26px;border-radius:8px;background:var(--accent);color:#fff;font:700 13px var(--sans)}}
.henry-block .secthead .dot{{background:var(--henry)}}
.secsub{{color:var(--muted);font-size:12px;line-height:1.5}} .secsub b{{color:var(--ink)}}
/* 立場分組帶 */
.stance-group{{margin:0 0 8px}}
.sband{{display:flex;align-items:center;gap:9px;margin:16px 0 11px;font:700 13px var(--sans);letter-spacing:.5px;color:var(--ink)}}
.sband .sdot{{width:11px;height:11px;border-radius:50%;background:var(--neu)}}
.sband em{{font-style:normal;font-family:var(--mono);font-size:11px;color:var(--muted);
  background:var(--panel);border:1px solid var(--hair);border-radius:20px;padding:1px 9px}}
.sband.s-bull .sdot{{background:var(--bull)}} .sband.s-bullmid .sdot{{background:var(--bullmid)}}
.sband.s-neu .sdot{{background:var(--neu)}} .sband.s-bearmid .sdot{{background:var(--bearmid)}}
.sband.s-bear .sdot{{background:var(--bear)}} .sband.s-watch .sdot{{background:var(--gd)}}
/* 卡片：兩欄、加大 */
.grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(452px,1fr));gap:16px}}
.card{{position:relative;background:var(--panel);border:1px solid var(--hair);border-radius:16px;
  padding:18px 20px 14px;box-shadow:var(--shadow);overflow:hidden}}
.card::before{{content:"";position:absolute;left:0;top:0;bottom:0;width:4px;background:var(--neu)}}
.card.s-bull::before{{background:var(--bull)}} .card.s-bullmid::before{{background:var(--bullmid)}}
.card.s-neu::before{{background:var(--neu)}} .card.s-bearmid::before{{background:var(--bearmid)}}
.card.s-bear::before{{background:var(--bear)}} .card.s-watch::before{{background:var(--gd)}}
.chead{{display:flex;align-items:center;gap:10px;margin-bottom:12px}}
.ident{{display:flex;align-items:baseline;gap:9px;margin-right:auto;min-width:0}}
.sym{{font-family:var(--serif);font-weight:700;font-size:21px;letter-spacing:.3px}}
.nm{{color:var(--muted);font-size:12px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}}
.badge{{padding:3px 12px;border-radius:20px;font-size:11.5px;font-weight:700;white-space:nowrap;flex:none}}
.badge.s-bull{{background:var(--beat-bg);color:var(--bull)}} .badge.s-bullmid{{background:#e7effb;color:var(--bullmid)}}
.badge.s-neu{{background:#efece5;color:var(--neu)}} .badge.s-bearmid{{background:#fbeee2;color:var(--bearmid)}}
.badge.s-bear{{background:var(--miss-bg);color:var(--bear)}} .badge.s-watch{{background:#f5edd8;color:var(--gd)}}
.qtag{{font-size:11px;color:var(--muted);letter-spacing:.2px;margin:-4px 0 12px;font-family:var(--mono)}}
/* 兩欄 body */
.cols{{display:grid;grid-template-columns:1fr 1fr;gap:0;border:1px solid var(--line);border-radius:12px;overflow:hidden}}
.col{{padding:13px 15px;display:flex;flex-direction:column;gap:13px}}
.col+.col{{border-left:1px solid var(--line)}}
.col-fwd{{background:var(--bg2)}}
.col-h{{font-size:10.5px;color:var(--muted);text-transform:uppercase;letter-spacing:.8px;font-weight:700;
  display:flex;align-items:center;gap:7px}}
.col-h .reported{{color:var(--beat);background:var(--beat-bg);border-radius:5px;padding:1px 6px;font-size:9px;letter-spacing:.3px}}
.metric{{}}
.mlabel{{font-size:10px;color:var(--muted);text-transform:uppercase;letter-spacing:.8px;font-weight:700;margin-bottom:4px}}
.mbig{{font-family:var(--serif);font-size:25px;font-weight:600;line-height:1.05;font-variant-numeric:tabular-nums;
  display:flex;align-items:baseline;gap:5px;flex-wrap:wrap}}
.mbig.sm{{font-size:19px}} .mbig.mut{{color:var(--muted);font-size:16px}}
.munit{{font-family:var(--sans);font-size:10.5px;color:var(--muted);font-weight:600}}
.diff{{display:flex;align-items:center;gap:8px;margin:7px 0 3px}}
.pill{{font-family:var(--sans);font-size:11.5px;font-weight:800;padding:2px 8px;border-radius:6px;letter-spacing:.2px;flex:none}}
.pill.beat{{background:var(--beat-bg);color:var(--beat)}} .pill.miss{{background:var(--miss-bg);color:var(--miss)}}
.cbar{{position:relative;height:6px;background:var(--panel);border:1px solid var(--line);border-radius:4px;flex:1;min-width:44px}}
.cmid{{position:absolute;left:50%;top:-1px;bottom:-1px;width:1.5px;background:var(--muted);opacity:.5}}
.cf{{position:absolute;top:0;bottom:0;border-radius:4px}}
.cf.beat{{background:var(--beat)}} .cf.miss{{background:var(--miss)}}
.cons{{font-size:11px;color:var(--muted);font-variant-numeric:tabular-nums}}
.cons b{{color:var(--ink2);font-weight:700}}
.guidbox{{background:#f7edd7;border-radius:9px;padding:9px 11px;margin-top:2px}}
.guidbox .gl{{display:block;font-size:9.5px;color:var(--gd);text-transform:uppercase;letter-spacing:.8px;font-weight:700;margin-bottom:3px}}
.guidbox .gv{{font-size:12.5px;color:#6b5416;font-weight:600;line-height:1.5}}
.gd{{font-size:12px;color:var(--ink2);line-height:1.68}}
.col-fwd .gd{{color:var(--ink2)}}
.stats{{display:grid;grid-template-columns:repeat(4,1fr);gap:1px;margin-top:13px;background:var(--line);
  border-radius:11px;overflow:hidden;border:1px solid var(--line)}}
.stat{{background:var(--panel);padding:9px 6px;text-align:center}}
.sl{{font-size:9.5px;color:var(--muted);letter-spacing:.5px;text-transform:uppercase;margin-bottom:3px}}
.sv{{font-family:var(--mono);font-size:13.5px;font-weight:600;font-variant-numeric:tabular-nums}}
.cview{{margin:14px 0 0;padding:12px 14px 12px 16px;background:var(--henry-bg);border-left:3px solid var(--henry);
  border-radius:0 10px 10px 0;font-size:12.5px;color:var(--ink2);line-height:1.72}}
.cview::before{{content:"“";font-family:var(--serif);color:var(--henry);font-size:20px;line-height:0;margin-right:2px;vertical-align:-4px;opacity:.5}}
footer{{color:var(--muted);font-size:11.5px;text-align:center;margin-top:30px;line-height:1.8;padding-top:18px;border-top:1px solid var(--line)}}
@media(max-width:520px){{
  body{{padding:16px 12px 44px}} h1{{font-size:25px}}
  .grid{{grid-template-columns:1fr;gap:13px}} .kpi{{min-width:100px;padding:11px 13px}}
  .cols{{grid-template-columns:1fr}} .col+.col{{border-left:0;border-top:1px solid var(--line)}}
}}
</style></head><body><div class="wrap">
{NAV}
<div class="masthead">
<h1>美股財報區</h1>
<p class="lede">兩位分析師的美股財報，<b>依多空立場分組</b>。每張卡兩欄：左＝最新季（實際 vs 共識的<b>預期差</b>），右＝下季<b>展望</b>（Allen 預估 vs 共識預期差＋公司指引）。<b>韭菜王 Allen</b>＝cover list 模型；<b>Henry</b>＝Pentimetrics「The Trace」券商彙整與 Bloomberg 共識。營收單位億、各檔原幣別。</p>
<div class="updated">UPDATED {updated}</div>
</div>
<div class="kpis">
 <div class="kpi"><div class="v">{n}</div><div class="l">Allen 覆蓋</div></div>
 <div class="kpi"><div class="v" style="color:var(--bull)">{bull}</div><div class="l">Allen 看多</div></div>
 <div class="kpi"><div class="v" style="color:var(--bear)">{bear}</div><div class="l">Allen 偏空</div></div>
 <div class="kpi"><div class="v" style="color:var(--henry)">{len(henry)}</div><div class="l">Henry 覆蓋</div></div>
</div>
<div class="bar">
 <input id="q" placeholder="🔍 搜尋代號／名稱…" oninput="flt()">
 <span class="chip on" data-s="all" onclick="pick(this)">全部立場</span>
 <span class="chip" data-s="看多" onclick="pick(this)">看多</span>
 <span class="chip" data-s="中性偏多" onclick="pick(this)">中性偏多</span>
 <span class="chip" data-s="中性" onclick="pick(this)">中性</span>
 <span class="chip" data-s="中性偏空" onclick="pick(this)">中性偏空</span>
 <span class="chip" data-s="偏空" onclick="pick(this)">偏空</span>
</div>

<section class="src-block">
<div class="secthead"><h2><span class="dot">A</span>韭菜王 Allen · Cover List</h2>
<span class="secsub">每季財報模型・依 Allen 立場分組（股價／PE 隨 GOOGLEFINANCE 日更）。</span></div>
{allen_groups}
</section>
{henry_sec}

<footer>Allen 數字取自 cover list（AMZN 用營業利益、PANW 用 FCF 故 EPS 留「—」；ASML／NOK 歐元、聯發科台幣，其餘美元；GOOG 2Q26 EPS 含未實現利得故 Beat% 偏高）。<br>
Henry 數字為 Pentimetrics「The Trace」原文摘錄、標出處期號／日期；立場依 Henry 原文評語判定；「市場預期」＝Bloomberg 共識，另附 buy side。數字一字不差、未明列者留空。<br>
Allen 區每日重跑更新；Henry 區為精選快照。© Evan 投資工作區</footer>
</div>
<script>
function pick(el){{document.querySelectorAll('.chip').forEach(c=>c.classList.remove('on'));el.classList.add('on');flt();}}
function flt(){{
  var q=document.getElementById('q').value.trim().toLowerCase();
  var s=document.querySelector('.chip.on').dataset.s;
  document.querySelectorAll('.card').forEach(function(c){{
    var okQ=(!q||c.dataset.s.indexOf(q)>=0);
    var okS=(s==='all'||c.dataset.stance===s);
    c.style.display=(okQ&&okS)?'':'none';
  }});
  document.querySelectorAll('.stance-group').forEach(function(g){{
    var any=[].some.call(g.querySelectorAll('.card'),x=>x.style.display!=='none');
    g.style.display=any?'':'none';
  }});
  document.querySelectorAll('.src-block').forEach(function(b){{
    var any=[].some.call(b.querySelectorAll('.card'),x=>x.style.display!=='none');
    b.style.display=any?'':'none';
  }});
}}
</script>
</body></html>"""


def main():
    wbd, wbf = load_wb()
    data = extract(wbd, wbf)
    henry = []
    hp = ROOT / "henry_data.json"
    if hp.exists():
        try:
            henry = json.loads(hp.read_text(encoding="utf-8"))
        except Exception as e:
            print("henry_data.json parse fail:", e)
    updated = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    (ROOT / "us_earnings.html").write_text(build_html(data, henry, updated), encoding="utf-8")
    print(f"wrote us_earnings.html  (Allen {len(data)} / Henry {len(henry)})")


if __name__ == "__main__":
    main()
